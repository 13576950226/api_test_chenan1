import logging
from logging import log
import openpyxl
import pymysql
import pytest
from openpyxl import load_workbook


class Link_Mysql:

    def link_myql(self, sql):
        connte = pymysql.connect(host="127.0.0.1",
                                 db="test",
                                 user="root",
                                 passwd="123456",
                                 port=3306,
                                 charset="utf8"
        )
        suc = connte.cursor()
        suc.execute(sql)
        sucprint = suc.fetchall()
        logging.error("aaaaa")
        print(sucprint)
        suc.close()
        connte.close()

    def red_or_wirt_execl(self):
        file = r"D:/test_case2.xlsx"
        workbook = load_workbook(file)
        worksheet = workbook.worksheets[0]
        for i in range(2, worksheet.max_row+1):
            worksheet.cell(i, 1).value = "99999999"
        workbook.save(r"D:/test_case2.xlsx")
        workbook.close()